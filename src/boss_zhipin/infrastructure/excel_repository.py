from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from boss_zhipin.models import Job


HEADERS = ["抓取日期", "首次发现", "是否新增", "职位", "公司", "薪资", "最低薪资(K)",
           "最高薪资(K)", "经验", "学历", "地点", "招聘者", "技能", "职位链接", "职位ID"]


class ExcelJobRepository:
    def __init__(self, path, legacy_path=None):
        self.path = Path(path)
        self.legacy_path = Path(legacy_path) if legacy_path else None

    def _workbook(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists() and self.legacy_path and self.legacy_path.exists():
            self.path.write_bytes(self.legacy_path.read_bytes())
        if self.path.exists():
            workbook = load_workbook(self.path)
            return workbook, workbook["职位"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "职位"
        sheet.append(HEADERS)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
        fill = PatternFill("solid", fgColor="16A085")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        return workbook, sheet

    def save(self, jobs):
        jobs = [job if isinstance(job, Job) else Job(**job) for job in jobs]
        workbook, sheet = self._workbook()
        existing = {str(sheet.cell(row, 15).value or ""): row for row in range(2, sheet.max_row + 1)}
        identities = {tuple(str(sheet.cell(row, col).value or "") for col in (4, 5, 6, 11)): row
                      for row in range(2, sheet.max_row + 1)}
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, 3).value = "否"
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_count = 0
        for job in jobs:
            identity = (job.job_name, job.company, job.salary, job.location)
            if job.job_id in existing:
                row, is_new = existing[job.job_id], "否"
            elif identity in identities:
                row, is_new = identities[identity], "否"
            else:
                row, is_new = sheet.max_row + 1, "是"
                new_count += 1
            first_seen = sheet.cell(row, 2).value or now
            values = [job.fetched_at, first_seen, is_new, job.job_name, job.company, job.salary,
                      job.salary_low, job.salary_high, job.experience, job.degree, job.location,
                      job.boss, job.skills, job.url, job.job_id]
            for column, value in enumerate(values, 1):
                sheet.cell(row, column).value = value
            if job.url:
                sheet.cell(row, 14).hyperlink = job.url
                sheet.cell(row, 14).style = "Hyperlink"
        widths = [20, 20, 10, 28, 24, 14, 14, 14, 14, 12, 22, 14, 32, 48, 38]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(sheet.max_row, 1)}"
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        workbook.save(temporary)
        temporary.replace(self.path)
        return new_count
