import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from boss_zhipin.infrastructure.excel_repository import ExcelJobRepository, HEADERS
from boss_zhipin.models import Job


class ExcelRepositoryTests(unittest.TestCase):
    def test_legacy_identity_is_migrated_without_duplicate(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "jobs.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "职位"
            sheet.append(HEADERS)
            sheet.append(["old", "first", "是", "前端开发", "示例公司", "15-25K", 15, 25,
                          "1-3年", "本科", "武汉", "招聘者", "Vue", "bad", "security-token"])
            workbook.save(output)
            job = Job("new", "前端开发", "示例公司", "15-25K", 15, 25, "1-3年", "本科",
                      "武汉", "招聘者", "Vue", "https://example/job", "short-id")

            count = ExcelJobRepository(output).save([job])
            result = load_workbook(output)["职位"]

        self.assertEqual(count, 0)
        self.assertEqual(result.max_row, 2)
        self.assertEqual(result.cell(2, 15).value, "short-id")

    def test_save_is_atomic_and_leaves_no_temporary_file(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "jobs.xlsx"
            job = Job("new", "前端", "公司", "15-20K", 15, 20, "不限", "本科", "武汉", "", "", "", "id")
            ExcelJobRepository(output).save([job])
            self.assertTrue(output.exists())
            self.assertFalse(output.with_suffix(".xlsx.tmp").exists())


if __name__ == "__main__":
    unittest.main()
