import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import boss_jobs
import browser_auth


class RuntimePathTests(unittest.TestCase):
    def test_runtime_paths_live_under_var(self):
        import runtime_paths

        expected = runtime_paths.WORK_DIR / "var"
        self.assertEqual(runtime_paths.RUNTIME_DIR, expected)
        self.assertEqual(runtime_paths.COOKIE_FILE, expected / "secrets" / "cookies.json")
        self.assertEqual(runtime_paths.CACHE_DIR, expected / "cache" / "security-js")
        self.assertEqual(runtime_paths.LOG_DIR, expected / "logs")
        self.assertEqual(runtime_paths.OUTPUT_DIR, expected / "outputs")
        self.assertEqual(runtime_paths.PROFILE_DIR, expected / "browser-profile")

    def test_cookie_loader_falls_back_to_legacy_file(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            new_cookie = root / "var" / "secrets" / "cookies.json"
            legacy_cookie = root / "cookies.json"
            legacy_cookie.write_text('{"zp_at": "legacy-token"}', encoding="utf-8")

            with patch.object(boss_jobs, "COOKIE_FILE", new_cookie), patch.object(
                boss_jobs, "LEGACY_COOKIE_FILE", legacy_cookie
            ), patch.object(boss_jobs, "COOKIE_TEXT_FILE", root / "var" / "secrets" / "cookies.txt"), patch.object(
                boss_jobs, "LEGACY_COOKIE_TEXT_FILE", root / "cookies.txt"
            ):
                self.assertEqual(boss_jobs.load_cookies(), {"zp_at": "legacy-token"})


class ParsingTests(unittest.TestCase):
    def test_salary_range(self):
        self.assertEqual(boss_jobs.parse_salary("15-25K·14薪"), (15.0, 25.0))

    def test_experience_upper_bound(self):
        self.assertEqual(boss_jobs.experience_max_years("1-3年"), 3)
        self.assertEqual(boss_jobs.experience_max_years("经验不限"), 0)
        self.assertEqual(boss_jobs.experience_max_years("3-5年"), 5)

    def test_encrypt_job_id_builds_detail_url(self):
        payload = {
            "zpData": {
                "jobList": [{
                    "salaryDesc": "15-25K",
                    "jobExperience": "1-3年",
                    "securityId": "not-a-job-id",
                    "encryptJobId": "d9b609bb4fa6d8a11H190t-0FFI~",
                    "jobName": "前端开发",
                    "brandName": "示例公司",
                    "cityName": "武汉",
                }]
            }
        }

        job = boss_jobs.extract_jobs(payload)[0]

        self.assertEqual(job["job_id"], "d9b609bb4fa6d8a11H190t-0FFI~")
        self.assertEqual(
            job["url"],
            "https://www.zhipin.com/job_detail/d9b609bb4fa6d8a11H190t-0FFI~.html",
        )


class WorkbookTests(unittest.TestCase):
    def test_legacy_security_id_row_is_migrated_without_duplicate(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "jobs.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "职位"
            sheet.append(boss_jobs.HEADERS)
            sheet.append([
                "old", "first", "是", "前端开发", "示例公司", "15-25K", 15, 25,
                "1-3年", "本科", "武汉", "招聘者", "Vue", "bad-url", "security-token",
            ])
            workbook.save(output)
            job = {
                "fetched_at": "new",
                "job_name": "前端开发",
                "company": "示例公司",
                "salary": "15-25K",
                "salary_low": 15,
                "salary_high": 25,
                "experience": "1-3年",
                "degree": "本科",
                "location": "武汉",
                "boss": "招聘者",
                "skills": "Vue",
                "url": "https://www.zhipin.com/job_detail/short-id.html",
                "job_id": "short-id",
            }

            with patch.object(boss_jobs, "EXCEL_FILE", output), patch.object(
                boss_jobs, "OUTPUT_DIR", Path(directory)
            ):
                new_count = boss_jobs.save_jobs([job])

            result = load_workbook(output)["职位"]
            self.assertEqual(result.max_row, 2)
            self.assertEqual(new_count, 0)
            self.assertEqual(result.cell(2, 15).value, "short-id")
            self.assertEqual(result.cell(2, 14).hyperlink.target, job["url"])


class BrowserAuthTests(unittest.TestCase):
    def test_login_requires_nonempty_zp_at(self):
        self.assertFalse(browser_auth.is_logged_in([]))
        self.assertFalse(browser_auth.is_logged_in([{"name": "zp_at", "value": ""}]))
        self.assertTrue(browser_auth.is_logged_in([{"name": "zp_at", "value": "token"}]))

    def test_only_zhipin_cookies_are_saved(self):
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "cookies.json"
            count = browser_auth.save_cookies([
                {"name": "zp_at", "value": "token", "domain": ".zhipin.com"},
                {"name": "wt2", "value": "visitor", "domain": "www.zhipin.com"},
                {"name": "other", "value": "secret", "domain": ".example.com"},
            ], destination)
            payload = __import__("json").loads(destination.read_text(encoding="utf-8"))

            self.assertEqual(count, 2)
            self.assertEqual({item["name"] for item in payload}, {"zp_at", "wt2"})
            self.assertEqual(destination.stat().st_mode & 0o777, 0o600)


if __name__ == "__main__":
    unittest.main()
