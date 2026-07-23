# pyright: reportMissingImports=false
import json
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.iv8_silent import import_iv8_silent
from utils.logger import logger
from runtime_paths import (
    CACHE_DIR,
    COOKIE_FILE,
    COOKIE_TEXT_FILE,
    EXCEL_FILE,
    LEGACY_COOKIE_FILE,
    LEGACY_COOKIE_TEXT_FILE,
    LEGACY_EXCEL_FILE,
    OUTPUT_DIR,
)
from boss_zhipin.domain.filtering import extract_jobs as extract_typed_jobs
from boss_zhipin.domain.filtering import experience_max_years
from boss_zhipin.domain.salary import parse_salary
from boss_zhipin.models import SearchCriteria


START_PAGE = 1
PAGE_COUNT = 5
PAGE_SIZE = 30
KEYWORD = "前端"
CITY_CODE = "101200100"  # 武汉
MIN_SALARY_K = 15
MAX_EXPERIENCE_YEARS = 3
REQUEST_INTERVAL_SECONDS = 1.5

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
PAGE_URL = f"https://www.zhipin.com/web/geek/jobs?query={urllib.parse.quote(KEYWORD)}&city={CITY_CODE}"
API_URL = "https://www.zhipin.com/wapi/zpgeek/search/joblist.json"

HEADERS = [
    "抓取日期", "首次发现", "是否新增", "职位", "公司", "薪资", "最低薪资(K)",
    "最高薪资(K)", "经验", "学历", "地点", "招聘者", "技能", "职位链接", "职位ID",
]


def apply_config(config):
    global START_PAGE, PAGE_COUNT, PAGE_SIZE, KEYWORD, CITY_CODE
    global MIN_SALARY_K, MAX_EXPERIENCE_YEARS, REQUEST_INTERVAL_SECONDS
    global PAGE_URL, CACHE_DIR, COOKIE_FILE, EXCEL_FILE, OUTPUT_DIR

    START_PAGE = config.search.start_page
    PAGE_COUNT = config.search.page_count
    PAGE_SIZE = config.search.page_size
    KEYWORD = config.search.keyword
    CITY_CODE = config.search.city_code
    MIN_SALARY_K = config.search.minimum_salary_k
    MAX_EXPERIENCE_YEARS = config.search.maximum_experience_years
    REQUEST_INTERVAL_SECONDS = config.request.interval_seconds
    PAGE_URL = f"https://www.zhipin.com/web/geek/jobs?query={urllib.parse.quote(KEYWORD)}&city={CITY_CODE}"
    CACHE_DIR = config.runtime.cache_dir
    COOKIE_FILE = config.runtime.cookie_file
    EXCEL_FILE = config.output.path
    OUTPUT_DIR = EXCEL_FILE.parent


def load_cookies():
    json_path = COOKIE_FILE if COOKIE_FILE.exists() else LEGACY_COOKIE_FILE
    text_path = COOKIE_TEXT_FILE if COOKIE_TEXT_FILE.exists() else LEGACY_COOKIE_TEXT_FILE
    if json_path.exists():
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            cookies = {str(k): str(v) for k, v in payload.items() if v is not None}
        elif isinstance(payload, list):
            cookies = {
                str(item["name"]): str(item["value"])
                for item in payload
                if isinstance(item, dict) and item.get("name") and item.get("value") is not None
            }
        else:
            raise ValueError("cookies.json 必须是 Cookie 对象或浏览器导出的 Cookie 数组")
    elif text_path.exists():
        cookie_text = text_path.read_text(encoding="utf-8").strip()
        cookies = {}
        for part in cookie_text.split(";"):
            if "=" not in part:
                continue
            name, value = part.strip().split("=", 1)
            if name:
                cookies[name] = value
    else:
        raise FileNotFoundError(
            f"缺少 {COOKIE_TEXT_FILE.name} 或 {COOKIE_FILE.name}。"
            "请运行 ./login.sh 完成登录并自动生成 Cookie。"
        )

    if not cookies or any("替换为" in value for value in cookies.values()):
        raise ValueError("Cookie 文件仍是示例或空内容，请填入你本人已登录会话的真实 Cookie")
    for name, value in cookies.items():
        try:
            name.encode("latin-1")
            value.encode("latin-1")
        except UnicodeEncodeError as exc:
            raise ValueError(f"Cookie {name!r} 含请求头无法编码的字符") from exc
    return cookies


def build_headers():
    return {
        "user-agent": UA,
        "accept": "application/json, text/plain, */*",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8",
        "content-type": "application/x-www-form-urlencoded",
        "origin": "https://www.zhipin.com",
        "referer": PAGE_URL,
        "x-requested-with": "XMLHttpRequest",
    }


def build_data(page):
    return {
        "scene": "1",
        "query": KEYWORD,
        "city": CITY_CODE,
        "experience": "",
        "degree": "",
        "industry": "",
        "scale": "",
        "salary": "",
        "jobType": "",
        "page": str(page),
        "pageSize": str(PAGE_SIZE),
    }


def replace_cookie(session, name, value):
    for cookie in list(session.cookies):
        if cookie.name == name:
            session.cookies.clear(domain=cookie.domain, path=cookie.path, name=cookie.name)
    session.cookies.set(name, str(value), domain=".zhipin.com", path="/")


def save_text(name, text):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    path = CACHE_DIR / name
    path.write_text(text, encoding="utf-8", errors="ignore")
    return path


def build_security_environment(security_url):
    parsed = urllib.parse.urlparse(security_url)
    return {
        "location": {
            "href": security_url,
            "origin": "https://www.zhipin.com",
            "protocol": "https:",
            "host": "www.zhipin.com",
            "hostname": "www.zhipin.com",
            "port": "",
            "pathname": parsed.path,
            "search": f"?{parsed.query}" if parsed.query else "",
            "hash": "",
        },
        "navigator": {
            "userAgent": UA,
            "platform": "MacIntel",
            "language": "zh-CN",
            "languages": ["zh-CN", "zh", "en"],
            "webdriver": False,
        },
        "screen": {
            "width": 1920, "height": 1080, "availWidth": 1920,
            "availHeight": 1040, "colorDepth": 24, "pixelDepth": 24,
        },
        "window": {"origin": "https://www.zhipin.com", "devicePixelRatio": 1},
        "canvas": {
            "fingerprint": {
                "toDataURL": {
                    "png": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAFgwJ/lQn8YQAAAABJRU5ErkJggg=="
                }
            }
        },
    }


def compute_stoken(session, challenge):
    zp = challenge.get("zpData") or {}
    seed, name, ts = zp["seed"], zp["name"], int(zp["ts"])
    js_url = f"https://www.zhipin.com/web/common/security-js/{name}.js"
    js_resp = session.get(js_url, headers={"user-agent": UA, "referer": PAGE_URL}, timeout=30)
    js_resp.raise_for_status()
    js_code = js_resp.text
    save_text(f"zhipin_security_{name}.js", js_code)

    security_url = "https://www.zhipin.com/web/common/security-check.html?" + urllib.parse.urlencode(
        {"seed": seed, "name": name, "ts": str(ts), "callbackUrl": "", "srcReferer": PAGE_URL}
    )
    html = f'<html><head><meta charset="utf-8"></head><body><script src="{js_url}"></script></body></html>'
    save_text("zhipin_security_check.html", html)
    snapshot = {"baseURL": security_url, "html": html, "headers": [], "resources": {js_url: js_code}}

    iv8 = import_iv8_silent()
    with iv8.JSContext(environment=build_security_environment(security_url), config={"timezone": "Asia/Shanghai"}) as ctx:
        ctx.expose(snapshot, "snapshot")
        ctx.eval("window.__iv8__.page.load(window.__iv8__.data.snapshot)")
        ctx.eval("window.__iv8__.eventLoop.sleep(100)")
        token = ctx.eval(f"encodeURIComponent((new window.ABC).z({json.dumps(seed)}, {ts}));")
    if not token:
        raise RuntimeError("iv8 未生成 __zp_stoken__")
    save_text("zhipin_stoken.txt", str(token))
    return str(token), seed, name, ts


def request_page(session, page):
    response = session.post(API_URL, headers=build_headers(), data=build_data(page), timeout=30)
    response.raise_for_status()
    payload = response.json()
    if payload.get("code") == 37 and payload.get("zpData"):
        logger.info("page={} 命中 code=37，刷新 __zp_stoken__", page)
        token, seed, name, ts = compute_stoken(session, payload)
        replace_cookie(session, "__zp_stoken__", token)
        replace_cookie(session, "__zp_sseed__", seed)
        replace_cookie(session, "__zp_sname__", name)
        replace_cookie(session, "__zp_sts__", ts)
        response = session.post(API_URL, headers=build_headers(), data=build_data(page), timeout=30)
        response.raise_for_status()
        payload = response.json()
    if payload.get("code") != 0:
        raise RuntimeError(f"BOSS API code={payload.get('code')}: {payload.get('message') or payload}")
    return payload


def extract_jobs(payload):
    criteria = SearchCriteria(KEYWORD, CITY_CODE, MIN_SALARY_K, MAX_EXPERIENCE_YEARS)
    return [job.to_dict() for job in extract_typed_jobs(payload, criteria)]

def load_or_create_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_FILE.exists() and LEGACY_EXCEL_FILE.exists():
        EXCEL_FILE.write_bytes(LEGACY_EXCEL_FILE.read_bytes())
    if EXCEL_FILE.exists():
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook["职位"]
        return workbook, sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "职位"
    sheet.append(HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    header_fill = PatternFill("solid", fgColor="16A085")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return workbook, sheet


def save_jobs(jobs):
    workbook, sheet = load_or_create_workbook()
    existing = {}
    existing_by_identity = {}
    for row in range(2, sheet.max_row + 1):
        existing[str(sheet.cell(row, 15).value or "")] = row
        identity = tuple(str(sheet.cell(row, col).value or "") for col in (4, 5, 6, 11))
        existing_by_identity[identity] = row

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 3).value = "否"

    new_count = 0
    for job in jobs:
        key = job["job_id"]
        identity = (job["job_name"], job["company"], job["salary"], job["location"])
        if key in existing:
            row = existing[key]
            first_seen = sheet.cell(row, 2).value
            is_new = "否"
        elif identity in existing_by_identity:
            # Migrate rows previously keyed by securityId to encryptJobId.
            row = existing_by_identity[identity]
            first_seen = sheet.cell(row, 2).value
            is_new = "否"
        else:
            row = sheet.max_row + 1
            first_seen = now
            is_new = "是"
            new_count += 1
        values = [
            job["fetched_at"], first_seen, is_new, job["job_name"], job["company"],
            job["salary"], job["salary_low"], job["salary_high"], job["experience"],
            job["degree"], job["location"], job["boss"], job["skills"], job["url"], key,
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col).value = value
        if job["url"]:
            sheet.cell(row, 14).hyperlink = job["url"]
            sheet.cell(row, 14).style = "Hyperlink"

    widths = [20, 20, 10, 28, 24, 14, 14, 14, 14, 12, 22, 14, 32, 48, 38]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(sheet.max_row, 1)}"
    workbook.save(EXCEL_FILE)
    return new_count


def main():
    cookies = load_cookies()
    session = requests.Session()
    session.cookies.update(cookies)
    logger.info("开始搜索：城市=武汉 关键词={} 最低薪资={}K 最大经验={}年", KEYWORD, MIN_SALARY_K, MAX_EXPERIENCE_YEARS)

    collected = {}
    failure = None
    for page in range(START_PAGE, START_PAGE + PAGE_COUNT):
        try:
            payload = request_page(session, page)
        except Exception as exc:
            failure = exc
            logger.info("page={} 请求失败，先保存已取得的 {} 个职位：{}", page, len(collected), exc)
            break
        page_jobs = extract_jobs(payload)
        for job in page_jobs:
            collected[job["job_id"]] = job
        logger.info("page={} 符合条件={} 累计={}", page, len(page_jobs), len(collected))
        time.sleep(REQUEST_INTERVAL_SECONDS)

    new_count = save_jobs(list(collected.values())) if collected else 0
    logger.info("保存完成：符合条件={} 新增={} Excel={}", len(collected), new_count, EXCEL_FILE)
    if failure:
        raise RuntimeError(f"任务未完成，已保存部分结果：{failure}") from failure


if __name__ == "__main__":
    main()
